#install wordpress
import os
if(os.name!='nt'):
    if(os.system('yum install httpd mariadb mariadb-server php-gd php-soap php-intl php-mysqlnd php-pdo php-pecl-zip php-fpm php-opcache php-curl php-zip php-xmlrpc wget')):
        if(os.system('systemctl start mariadb')):
            if(os.system('mysql_secure_installation')):
                
            else:
                print('check maria secures')
        else:
            print('maria db not started')
    else:
        print(os.error)
'''
from  openpyxl import load_workbook
wb=load_workbook('asd.xlsx')
sh=wb['titlqena']
print(sh['A1'].value)
sh['A1'].value='test'
wb.save('asd.xlsx')

from openpyxl import Workbook
wb=Workbook()
wb.create_sheet('titlqena',0)
sheet=wb['titlqena']
sheet['A1']='test create sheet'
wb.save('asd.xlsx')

from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
print(sheet.title)
sheet.title='demo'
celln=sheet.cell(row=1,column=1)
celln.value='qena'
print(celln.value)
sheet['D2']='cyber'
wb.save('test.xlsx')
'''