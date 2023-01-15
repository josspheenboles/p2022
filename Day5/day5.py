from openpyxl import load_workbook
wb=load_workbook('asd.xlsx')
sh=wb['sheet1']
print(sh['A1'].value)
sh['c1']='qena'
wb.save('asd.xlsx')
'''
from openpyxl import Workbook
wb=Workbook()#clear file
#sh=wb.active
wb.create_sheet('sheet1')
sh=wb['sheet1']
sh['A1']='qena'
sh['B1']='sys admin'
sh['A2']='q4 2022-2023'
sh['A3']=1.1
wb.save('asd.xlsx')


import socket
import sys
#f=open('asd.txt','r')
#ip=f.read()
#f.close()
ips=sys.argv[1:]
ip=sys.argv[1]
hostname=socket.gethostbyname(ip)
try:
    #port 1 to 65535
    for port in range(70,81):
        print(port)
        #connect to hostname 1--->80
        s=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
        socket.setdefaulttimeout(.5)
        result=s.connect_ex((hostname,port))#(target host,port)
        if(result==0):
            print('port number',port,'is opened')
        s.close()
except KeyboardInterrupt:
    print('exsitting program')
except socket.gaierror:
    print('hostname cannt resolved')
except socket.error:
    print('server not responding')
except:
    print('error')

import re
#xx@xx.xx
pattern=r'[\w.+-]+@[\w-]+\.[\w.-]+'
str1='wkelke asd@yahoo.com fkjdlkjds asd@gmail.com'
print(re.search(pattern,str1))
print(re.match(pattern,str1))
print(re.fullmatch(pattern,str1))


import os
os.chdir('H:\\')
print(os.system('dir'))
print(os.name)
os.system('su admin1')

if(os.name=='nt'):
    os.system('dir')
else:
    os.system('ls -l')
    if(os.getlogin()=='root'):
        os.system('user add test2')
    else:
        print('you must login as root')

import math#standred module
print(math.pow(2,4))
print(math.pi)



def dosum(x,y):
    print(x+y)
def dosum(x):
    print(x+10)
from myexperiance.files.textfileoperation import read as textread
from myexperiance.files.excelfileoperaion import read as excelread
textread()
excelread('asd.txt')





from myexperiance.files.textfileoperation import *
print(readmode)
print(readfromtextfile('asas.txt',))

import myexperiance.files.textfileoperation as  qenamodule1
#print(myexperiance.files.textfileoperation.readmode)
print(qenamodule1.modes)

from myexperiance.files.textfileoperation import modes
#by name not moudlename.blockname
print(modes)


import textfileoperation
print(textfileoperation.readfromtextfile('asd.txt','read'))
#textfileoperation.writetotextfile('asd.txt')
print(textfileoperation.modes)

from textfileoperation import readfromtextfile


textfileoperation.writetotextfile('test.txt','test')
textfileoperation.writetotextfile('asd.txt',['test\n','test2\n'])
print(textfileoperation.readfromtextfile('asd.txt','read'))
print(readfromtextfile('asd.txt','read',2))
print(readfromtextfile('asd.txt','line'))
print(readfromtextfile('asd.txt','lines'))
'''