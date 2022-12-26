from openpyxl import workbook
wb=workbook.Workbook()
sh=wb.active
#read log files access apache centos8 --->simple file
file=open('/var/logs/apach2/access.log','r')
lines=file.readlines()
for line in lines:
    if('26/12/2022' in line):
        sh['A1']=line
        #write
file.close()
wb.save('test.xlsx')
'''
from openpyxl import load_workbook
wb=load_workbook('asd.xlsx')
#print(type(wb))
sh=wb.active
sh['A1']='aswan2'
sh['B1']='sys admin'
sh['C1']=2022
wb.save('asd.xlsx')

#create excel 
from openpyxl import workbook
#create excel file
wb=workbook.Workbook()
wb.create_sheet('test')
sh=wb[wb.sheetnames[1]]
print(wb.sheetnames)
sh['B1']='asd'

wb.save('asd.xlsx')
'''